#!/usr/bin/env python3
"""
check_urpl.py — porównuje strains.json z oficjalną Listą Surowców Farmaceutycznych (URPL/eZdrowie).

Co robi:
  1. Pobiera aktualny wykaz surowców farmaceutycznych (susz konopny i in.) z rejestru URPL (plik XLSX).
  2. Filtruje pozycje związane z konopiami (Cannabis flos / kwiatostan).
  3. Dopasowuje je do Twojej bazy strains.json po nazwie odmiany + producencie.
  4. Wypisuje raport:
       - NOWE w URPL, których nie ma w strains.json  -> do dodania
       - SĄ w strains.json, ale ZNIKNĘŁY z URPL       -> prawdopodobnie wycofane
       - RÓŻNICE w THC/CBD między URPL a strains.json -> do weryfikacji

Czego NIE robi:
  - Nie pobiera i nie aktualizuje profili terpenowych (URPL ich nie publikuje —
    to trzeba weryfikować ręcznie/przez Claude względem BudCare.pl i innych źródeł).
  - Nie modyfikuje strains.json automatycznie — tylko raportuje różnice.
    Decyzję o edycji podejmujesz Ty (dane kliniczne = zero cichych automatycznych zmian).

Wymagania:
    pip install requests openpyxl

Użycie:
    python3 check_urpl.py
    python3 check_urpl.py --strains-file strains.json
    python3 check_urpl.py --save-raw urpl_raw.xlsx   # zachowaj pobrany plik do ręcznej inspekcji

Uwaga:
  Adres endpointu i nazwy kolumn w pliku URPL mogą się zmienić — jeśli skrypt
  nie znajdzie oczekiwanych kolumn, wypisze listę kolumn z pliku, żebyś mógł
  poprawić mapowanie w sekcji COLUMN HINTS poniżej. Nie testowałem tego skryptu
  na żywym pobraniu (środowisko, w którym go pisałem, nie miało dostępu do sieci) —
  przy pierwszym uruchomieniu sprawdź uważnie output.
"""

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Brakuje pakietu 'requests'. Zainstaluj: pip install requests openpyxl")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Brakuje pakietu 'openpyxl'. Zainstaluj: pip install requests openpyxl")

URPL_XLSX_URL = "https://rejestry.ezdrowie.gov.pl/api/rpl/medicinal-products/pharmaceutical-raw-materials/get-xlsx"

# ---------- COLUMN HINTS ----------
# Fragmenty nazw kolumn (case-insensitive, bez polskich znaków) po których szukamy
# odpowiednich pól w arkuszu URPL. Jeśli URPL zmieni nazewnictwo, dopisz tu warianty.
COL_HINTS = {
    "name": ["nazwa", "nazwa surowca", "nazwa produktu"],
    "producer": ["podmiot", "wytworca", "producent", "posiadacz"],
    "thc": ["thc"],
    "cbd": ["cbd"],
    "status": ["status", "wazno", "decyzj"],
}

CANNABIS_KEYWORDS = ["cannabis", "konopi", "kwiatostan", "flos"]


def strip_pl(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def find_column(header_row, hints):
    for idx, cell in enumerate(header_row):
        cell_norm = strip_pl(cell)
        for h in hints:
            if h in cell_norm:
                return idx
    return None


def download_urpl_xlsx(save_raw=None):
    print(f"Pobieram rejestr URPL: {URPL_XLSX_URL}")
    resp = requests.get(URPL_XLSX_URL, timeout=60)
    resp.raise_for_status()
    if save_raw:
        Path(save_raw).write_bytes(resp.content)
        print(f"Zapisano surowy plik: {save_raw}")
    import io
    return load_workbook(io.BytesIO(resp.content), data_only=True)


def parse_urpl(wb):
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Pusty arkusz URPL — sprawdź --save-raw i otwórz plik ręcznie.")

    header = rows[0]
    col_name = find_column(header, COL_HINTS["name"])
    col_producer = find_column(header, COL_HINTS["producer"])
    col_thc = find_column(header, COL_HINTS["thc"])
    col_cbd = find_column(header, COL_HINTS["cbd"])

    if col_name is None:
        print("\n⚠️  Nie znalazłem kolumny z nazwą surowca automatycznie.")
        print("Kolumny znalezione w pliku:")
        for i, c in enumerate(header):
            print(f"  [{i}] {c}")
        sys.exit("Popraw COL_HINTS w skrypcie i uruchom ponownie.")

    entries = []
    for row in rows[1:]:
        name = row[col_name] if col_name is not None else None
        if not name:
            continue
        name_str = str(name)
        if not any(k in strip_pl(name_str) for k in CANNABIS_KEYWORDS):
            continue
        entries.append({
            "raw_name": name_str.strip(),
            "producer": str(row[col_producer]).strip() if col_producer is not None and row[col_producer] else "",
            "thc": str(row[col_thc]).strip() if col_thc is not None and row[col_thc] else "",
            "cbd": str(row[col_cbd]).strip() if col_cbd is not None and row[col_cbd] else "",
        })
    return entries


def load_local_strains(path):
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return data.get("strains", data if isinstance(data, list) else [])


def normalize_pct(s):
    """Wyciąga liczbę z '20%', '≤1%', '18-22%' (bierze pierwszą liczbę) do porównań przybliżonych."""
    if not s:
        return None
    m = re.search(r"[\d,.]+", str(s))
    if not m:
        return None
    return float(m.group(0).replace(",", "."))


def strain_name_in_urpl_text(strain_name, urpl_raw_name):
    """URPL zwykle zapisuje pełną frazę typu 'Cannabis flos Aurora 29 Sourdough' —
    więc dopasowujemy przez sprawdzenie, czy znormalizowana nazwa odmiany
    (lub jej pierwszy człon, dla nazw wieloczłonowych typu 'Blue Monkey / Mac Monkey')
    występuje jako podciąg w pełnej nazwie z rejestru."""
    variants = [strain_name] + [v.strip() for v in re.split(r"[/]", strain_name)]
    urpl_norm = strip_pl(urpl_raw_name)
    for v in variants:
        v_norm = strip_pl(v)
        if v_norm and v_norm in urpl_norm:
            return True
    return False


def fuzzy_key(name, producer):
    return strip_pl(name) + "|" + strip_pl(producer).split()[0] if producer else strip_pl(name)


def main():
    ap = argparse.ArgumentParser(description="Sprawdź strains.json względem rejestru URPL")
    ap.add_argument("--strains-file", default="strains.json")
    ap.add_argument("--save-raw", default=None, help="zapisz pobrany plik xlsx pod tą ścieżką")
    args = ap.parse_args()

    local = load_local_strains(args.strains_file)

    wb = download_urpl_xlsx(save_raw=args.save_raw)
    urpl_entries = parse_urpl(wb)

    if not urpl_entries:
        print("Nie znaleziono pozycji zawierających słowa kluczowe (cannabis/konopi/flos)."
              " Sprawdź --save-raw i nazewnictwo w pliku ręcznie.")
        return

    print(f"\nZnaleziono {len(urpl_entries)} pozycji konopnych w URPL, "
          f"{len(local)} odmian w {args.strains_file}.\n")

    urpl_keys_seen = set()
    new_in_urpl, thc_mismatch = [], []
    matched_local_names = set()

    for e in urpl_entries:
        # znajdź dopasowanie w lokalnej bazie po zawieraniu się nazwy w pełnej frazie URPL
        candidates = [s for s in local if strain_name_in_urpl_text(s["name"], e["raw_name"])]
        # jeśli producent znany, zawęź po pierwszym słowie producenta (odsiewa przypadkowe trafienia typu "MAC 1" u kilku firm)
        if e["producer"]:
            prod_first_word = strip_pl(e["producer"]).split()[0] if strip_pl(e["producer"]) else ""
            narrowed = [s for s in candidates if prod_first_word and prod_first_word in strip_pl(s.get("producer", ""))]
            if narrowed:
                candidates = narrowed

        if not candidates:
            new_in_urpl.append(e)
            continue

        local_match = candidates[0]
        matched_local_names.add(fuzzy_key(local_match["name"], local_match.get("producer", "")))

        urpl_thc = normalize_pct(e["thc"])
        local_thc = normalize_pct(local_match.get("thc"))
        if urpl_thc is not None and local_thc is not None and abs(urpl_thc - local_thc) >= 1.0:
            thc_mismatch.append((local_match["name"], local_match.get("producer", ""), local_thc, urpl_thc))

    missing_from_urpl = []
    for s in local:
        key = fuzzy_key(s["name"], s.get("producer", ""))
        if key not in matched_local_names:
            missing_from_urpl.append(s)

    print("=" * 60)
    print(f"NOWE W URPL, brak w {args.strains_file} ({len(new_in_urpl)}):")
    print("=" * 60)
    if not new_in_urpl:
        print("  (brak)")
    for e in new_in_urpl:
        print(f"  + {e['raw_name']}  [{e['producer']}]  THC {e['thc']}  CBD {e['cbd']}")

    print()
    print("=" * 60)
    print(f"W {args.strains_file}, ale NIE MA już w URPL — prawdopodobnie wycofane ({len(missing_from_urpl)}):")
    print("=" * 60)
    if not missing_from_urpl:
        print("  (brak)")
    for s in missing_from_urpl:
        already_flagged = "WYCOFANA" in str(s.get("status", ""))
        flag = "  (już oznaczone jako wycofane)" if already_flagged else "  ⚠️ SPRAWDŹ — nieoznaczone w bazie"
        print(f"  - {s['name']}  [{s.get('producer','')}]{flag}")

    print()
    print("=" * 60)
    print(f"RÓŻNICE THC ≥1pp między URPL a {args.strains_file} ({len(thc_mismatch)}):")
    print("=" * 60)
    if not thc_mismatch:
        print("  (brak)")
    for name, producer, local_thc, urpl_thc in thc_mismatch:
        print(f"  ~ {name} [{producer}]: baza={local_thc}%  URPL={urpl_thc}%")

    print()
    print("Pamiętaj: profile terpenowe nie są w tym rejestrze — do ich weryfikacji "
          "użyj BudCare.pl ręcznie albo poproś o to w rozmowie z Claude.")


if __name__ == "__main__":
    main()
